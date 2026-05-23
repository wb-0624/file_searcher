# -*- coding: utf-8 -*-
"""
处理 .xlsx 文件，提取所有单元格的文本
"""
import logging
from openpyxl import load_workbook

logger = logging.getLogger(__name__)

def extract_xlsx_text(file_path):
    """
    提取xlsx文件所有单元格文本（字符串类型）。
    :param file_path: xlsx文件路径
    :return: 提取的文本，失败返回空字符串
    """
    logger.debug(f"提取xlsx文本: {file_path}")
    try:
        wb = load_workbook(file_path, data_only=True)
        texts = []
        for sheet in wb.worksheets:
            for row in sheet.iter_rows(values_only=True):
                for cell in row:
                    if cell and isinstance(cell, str):
                        texts.append(cell)
        text = '\n'.join(texts)
        logger.debug(f"成功提取 {len(texts)} 个单元格，总字符数: {len(text)}")
        return text
    except Exception as e:
        logger.error(f"读取xlsx失败 {file_path}: {e}")
        return ""